import re
import math
import pandas
import openpyxl
import pandas as pd
from openpyxl import Workbook
from Class import Block	

# Initialize variable from Excel
Excel_file = '../data/System_Money.xlsx'
Block_Data = pandas.read_excel(Excel_file, sheet_name = 'blocks')
Expend_Data = pandas.read_excel(Excel_file, sheet_name = 'expend')
Allocation_Data = pandas.read_excel(Excel_file, sheet_name = 'allocation')

# Initialize list obj blocks from Excel
List_Block = []

for index_row, row in Block_Data.iterrows():
	List_Block.append(Block(row['Name']))

# Extraction expend data from Excel
for index_row, row in Expend_Data.iterrows():
	block = [block for block in List_Block if block.name == row['Block']][0]
	if math.isnan(row['Val']):
		raise Exception("In Sheet Expend Val Not Empty")
	block.add_record(row['Comment'], row['Val'], row['Date'])

# Extraction allocation data from Excel
for index_row, row in Allocation_Data.iterrows():
	block = [block for block in List_Block if block.name == row['Block']][0]
	if math.isnan(row['Val']):
		raise Exception("In Sheet Allocation Val Not Empty")
	block.add_allocation(row['Val'], row['Date'])

# Insert Analyze date to Excel
Excel_file = '../data/Analyze_file.xlsx'

book = Workbook()

for block in List_Block:
	sheet = book.create_sheet(block.name)	
	for month in block.list_month:
		sheet.append(["month", month.name, month.month])
		sheet.append(["Comment", "Val", "Date"])
		for record in month.list_record:
			sheet.append([record['comment'], record['val'], record['date']])
	
		sheet.append(["", "", ""])
	
		sheet.append(["Allocation", month.allocation, "***"])
		sheet.append(["Expend", month.expend, "***"])
		sheet.append(["Remaining", month.get_remaining(), "***"])

		sheet.append(["", "", ""])
		sheet.append(["", "", ""])
	

# Initialize result page in Excel
sheet = book.create_sheet('result')

sheet.append(['Block', 'Allocation', 'Expend', 'remaining'])

sum_allocaton = 0
sum_expend = 0
sum_remaining = 0

for block in List_Block:
	sheet.append([block.name, block.get_total_alocation(), block.get_total_expend(), block.get_total_remaining()])
	sum_allocaton += block.get_total_alocation()
	sum_expend += block.get_total_expend()
	sum_remaining += block.get_total_remaining()

sheet.append(['', '', '', ''])
sheet.append(['sum', sum_allocaton, sum_expend, sum_remaining])

# Del sheet and save Excel file
del book['Sheet']
book.save(Excel_file)
