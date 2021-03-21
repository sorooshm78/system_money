import pandas
import openpyxl
import pandas as pd
from openpyxl import Workbook
	
# Class Block
class Block:
	def __init__(self, name):
		self.name = name
		self.allocation = 0
		self.expend = 0 
		self.list_record = []

	def print(self):
		print("Name:" + self.name +
		"\t allo:" +  str(self.allocation) +
		"\t expend:" + str(self.expend) +
		"\t remaining:" + str(self.get_remaining()))

		print("list records:")
		for record in self.list_record:	
			print("comment:" + record['comment'] +
			"\t val:" +  str(record['val']) +
			"\t data:" + record['data'])

	def add_expend(self, val):
		self.expend += val 

	def add_allocation(self, val):
		self.allocation += val 

	def get_remaining(self):
		return self.allocation - self.expend
	
	def add_record(self, comment, val, data):
		self.list_record.append({"comment" : comment,
								 "val" : val,
								 "data" : data})
	def __key_by_val(dict):
		return dict['val']

	def __key_by_data(dict):
		return dict['data']

	def sort_record_by_val:
		list_record.sort(key=key_by_val)
		
	def sort_record_by_data:
		list_record.sort(key=key_by_data)

# Initialize variable from Excel
Block_Data = pandas.read_excel('System_Money.xlsx', sheet_name = 'blocks')
Expend_Data = pandas.read_excel('System_Money.xlsx', sheet_name = 'expend')
Allocation_Data = pandas.read_excel('System_Money.xlsx', sheet_name = 'allocation')

# Initialize list obj blocks from Excel
List_Block = []

for index_row, row in Block_Data.iterrows():
	List_Block.append(Block(row['Name']))

# function
def find_obj(name):
	for block in List_Block:
		if name	== block.name:
			return block

# Extraction expend data from Excel
for index_row, row in Expend_Data.iterrows():
	find_obj(row['Block']).add_expend(row['Val'])
	find_obj(row['Block']).add_record(row['Comment'], row['Val'], row['Data'])

# Extraction allocation data from Excel
for index_row, row in Allocation_Data.iterrows():
	find_obj(row['Block']).add_allocation(row['Val'])

# Insert Analyze date to Excel
Excel_file = 'Analyze_file.xlsx'

book = Workbook()

for block in List_Block:
	sheet = book.create_sheet(block.name)	
	sheet.append(["Comment", "Val", "Data"])
	for record in block.list_record:
		sheet.append([record['comment'], record['val'], record['data']])

	sheet.append(["", "", ""])
	sheet.append(["Allocation", block.allocation, "***"])
	sheet.append(["Expend", block.expend, "***"])
	sheet.append(["Remaining", block.get_remaining(), "***"])

# Initialize result page in Excel
sheet = book.create_sheet('result')

sheet.append(['Block', 'Allocation', 'Expend', 'remaining'])

sum_allocaton = 0
sum_expend = 0
sum_remaining = 0

for block in List_Block:
	sheet.append([block.name, block.allocation, block.expend, block.get_remaining()])
	sum_allocaton += block.allocation
	sum_expend += block.expend
	sum_remaining += block.get_remaining()

sheet.append(['', '', '', ''])
sheet.append(['sum', sum_allocaton, sum_expend, sum_remaining])

# Del sheet and save Excel file
del book['Sheet']
book.save(Excel_file)
"""
